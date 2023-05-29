# encoding:utf-8
import json
import re
import time
import numpy as np
import openai
import openai.error
import requests
from bot.bot import Bot
from bot.chatgpt.chat_gpt_session import ChatGPTSession
from bot.openai.open_ai_image import OpenAIImage
from bot.session_manager import SessionManager
from bridge.context import ContextType
from bridge.reply import Reply, ReplyType
from common.log import logger
from common.token_bucket import TokenBucket
from config import conf, load_config
import openpyxl

def xingpan(socket_birth_data: dict):
    header = {"Content-Type": "application/json;charset=UTF-8"}
    # socket_required_transform = openai.ChatCompletion.create(
    #     model="gpt-3.5-turbo",
    #     temperature=0.5,
    #     max_tokens=512,
    #     messages=[
    #         {"role": "system","content": '将我的输入转换成以下格式输出：{"birthday": "1994-09-0115 5:30","latitude": "30.60","longitude": "114.30"}，其中城市转化为经纬度，没有输入时间的话默认为中午12:00'},
    #         {"role": "user", "content": birth_data},
    #     ]
    # )
    # potential_socket_data = socket_required_transform['choices'][0]['message']['content'] #将用户输入的生日转化成接口所需格式
    # if "latitude" in potential_socket_data:                 #判断是否正确返回了接口所需格式
    # socket_birth_data = json.loads(potential_socket_data)
    natal_url = 'http://www.xingpan.vip/astrology/chart/natal'
    natal_required_datas = {
        "access_token": "26254c620946fc59352afd441b61fb66",
        "h_sys": "P",
        "planets": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "t","H"],
        # "planets": ["0", "1", "2"],
        "tz": "8.00",
        list(socket_birth_data.keys())[0]: list(socket_birth_data.values())[0],
        list(socket_birth_data.keys())[1]: list(socket_birth_data.values())[1],
        list(socket_birth_data.keys())[2]: list(socket_birth_data.values())[2],
        'is_corpus': 1,
        'phase': {'0': 0.5, '180': 6, '120': 6, '90': 6, '60': 6}
       # 'phase': {'0': 0.5, '180': 6, '120': 6, '90': 6, '60': 6, '30': 2, '36': 2, '45': 2, '72': 2, '135': 0.5,
        #          '144': 2, '150': 2}
    }
    returned_natal = requests.post(natal_url, headers=header, data=json.dumps(natal_required_datas)).json()  # 得到星盘数据
    house_and_sign = []  # 宫位和落座
    xiangwei = []  # 相位
    natal_text = '星盘是：'
    for i in range(len(returned_natal['data']['planet'])):
        house_and_sign.append({'type': 4,
                               'planet_id': returned_natal['data']['planet'][i]['code_name'],
                               'house_id': returned_natal['data']['planet'][i]['house_id'], })
        house_and_sign.append({'type': 5,
                               'planet_id': returned_natal['data']['planet'][i]['code_name'],
                               'sign_id': returned_natal['data']['planet'][i]['sign']['sign_id'], })  # 从星盘里获取宫位和落座的数据
        # print(r['data']['planet'][j]['planet_allow_degree'])
        # print(len(r['data']['planet'][j]['planet_allow_degree']))
        for j in range(len(returned_natal['data']['planet'][i]['planet_allow_degree'])):
            xiangwei.append({'type': 6,
                             'planet_id1': returned_natal['data']['planet'][i]['code_name'],
                             'planet_id2': returned_natal['data']['planet'][i]['planet_allow_degree'][j]['code_name'],
                             'degree': returned_natal['data']['planet'][i]['planet_allow_degree'][j][
                                 'allow']})  # 从星盘里获取相位的数据
    socket_required_hs = {'chartType': 'natal',
                          "access_token": "26254c620946fc59352afd441b61fb66",
                          "fallInto": json.dumps(house_and_sign)}  # 转化宫位和落座数据以传入接口
    socket_required_xw = {'chartType': 'natal',
                          "access_token": "26254c620946fc59352afd441b61fb66",
                          "fallInto": json.dumps(xiangwei)}  # 转化相位数据以传入接口
    yulianurl = 'http://www.xingpan.vip/astrology/corpusconstellation/getlist'
    hs_text = requests.post(yulianurl, headers=header, data=json.dumps(socket_required_hs)).json()  # 宫位和落座语料
    for h in range(len(hs_text['data'])):
        natal_text += hs_text['data'][h]['title'] + '，'
    xw_text = requests.post(yulianurl, headers=header, data=json.dumps(socket_required_xw)).json()  # 相位语料
    for x in range(len(xw_text['data'])):
        natal_text += xw_text['data'][x]['title'] + '，'
    natal_text = natal_text[:-1] + '。'  # 结尾改成句号
    return returned_natal, natal_text
    # else: return birth_data #如果没返回接口所需格式将直接返回输入原文，不执行任何操作


def get_title(type: int, id1:str, id2, agree=999):
    house = ['', '1宫', '2宫', '3宫', '4宫', '5宫', '6宫', '7宫', '8宫', '9宫', '10宫', '11宫', '12宫']
    sign = ['', '白羊座', '金牛座', '双子座', '巨蟹座', '狮子座', '处女座', '天秤座', '天蝎座', '射手座', '摩羯座',
            '水瓶座', '双鱼座']
    planet = {'0': '太阳', '1': '月亮', '2': '水星', '3': '金星', '4': '火星', '5': '木星', '6': '土星', '7': '天王',
              '8': '海王', '9': '冥王', 't': '北交', 'H': '婚神'}
    xiagnwei = {0: '合', 180: '冲', 120: '拱', 90: '刑', 60: '六合'}
    title = ''
    title += planet[id1]
    if type == 4:
        title += house[id2]  # 行星落宫位
    elif type == 5:
        title += sign[id2]  # 行星落星座
    elif type == 6:
        title = title + xiagnwei[agree] + planet[id2]  # 行星相位
    return title

# OpenAI对话模型API (可用)
class ChatGPTBot(Bot, OpenAIImage):
    def __init__(self):
        super().__init__()
        # set the default api_key
        openai.api_key = conf().get("open_ai_api_key")
        if conf().get("open_ai_api_base"):
            openai.api_base = conf().get("open_ai_api_base")
        proxy = conf().get("proxy")
        if proxy:
            openai.proxy = proxy
        if conf().get("rate_limit_chatgpt"):
            self.tb4chatgpt = TokenBucket(conf().get("rate_limit_chatgpt", 20))

        self.sessions = SessionManager(
            ChatGPTSession, model=conf().get("model") or "gpt-3.5-turbo"
        )
        self.args = {
            "model": conf().get("model") or "gpt-3.5-turbo",  # 对话模型的名称
            "temperature": conf().get("temperature", 0.7),  # 值在[0,1]之间，越大表示回复越具有不确定性
            "max_tokens": 512,  # 回复最大的字符数
            "top_p": 1,
            "frequency_penalty": conf().get(
                "frequency_penalty", 0.0
            ),  # [-2,2]之间，该值越大则更倾向于产生不同的内容
            "presence_penalty": conf().get(
                "presence_penalty", 0.0
            ),  # [-2,2]之间，该值越大则更倾向于产生不同的内容
            "request_timeout": conf().get(
                "request_timeout", None
            ),  # 请求超时时间，openai接口默认设置为600，对于难问题一般需要较长时间
            "timeout": conf().get("request_timeout", None),  # 重试超时时间，在这个时间内，将会自动重试
        }

    def reply(self, query, context=None):
        # acquire reply content
        if context.type == ContextType.TEXT:
            logger.info("[CHATGPT] query={}".format(query))

            session_id = context["session_id"]
            reply = None
            clear_memory_commands = conf().get("clear_memory_commands", ["#清除记忆"])
            if query in clear_memory_commands:
                self.sessions.clear_session(session_id)
                reply = Reply(ReplyType.INFO, "记忆已清除")
            elif query == "#清除所有":
                self.sessions.clear_all_session()
                reply = Reply(ReplyType.INFO, "所有人记忆已清除")
            elif query == "#更新配置":
                load_config()
                reply = Reply(ReplyType.INFO, "配置已更新")
            if reply:
                return reply
            session = self.sessions.session_query(query, session_id)
            logger.debug("[CHATGPT] session query={}".format(session.messages))

            api_key = context.get("openai_api_key")

            # if context.get('stream'):
            #     # reply in stream
            #     return self.reply_text_stream(query, new_query, session_id)

            reply_content = self.reply_text(session, api_key)
            logger.debug(
                "[CHATGPT] new_query={}, session_id={}, reply_cont={}, completion_tokens={}".format(
                    session.messages,
                    session_id,
                    reply_content["content"],
                    reply_content["completion_tokens"],
                )
            )
            if (
                    reply_content["completion_tokens"] == 0
                    and len(reply_content["content"]) > 0
            ):
                reply = Reply(ReplyType.ERROR, reply_content["content"])
            elif reply_content["completion_tokens"] > 0:
                self.sessions.session_reply(
                    reply_content["content"], session_id, reply_content["total_tokens"]
                )
                reply = Reply(ReplyType.TEXT, reply_content["content"])
            else:
                reply = Reply(ReplyType.ERROR, reply_content["content"])
                logger.debug("[CHATGPT] reply {} used 0 tokens.".format(reply_content))
            return reply

        elif context.type == ContextType.IMAGE_CREATE:
            ok, retstring = self.create_img(query, 0)
            reply = None
            if ok:
                reply = Reply(ReplyType.IMAGE_URL, retstring)
            else:
                reply = Reply(ReplyType.ERROR, retstring)
            return reply
        else:
            reply = Reply(ReplyType.ERROR, "Bot不支持处理{}类型的消息".format(context.type))
            return reply

    def reply_text(self, session: ChatGPTSession, api_key=None, retry_count=0) -> dict:
        """
        call openai's ChatCompletion to get the answer
        :param session: a conversation session
        :param session_id: session id
        :param retry_count: retry count
        :return: {}
        """
        try:
            if conf().get("rate_limit_chatgpt") and not self.tb4chatgpt.get_token():
                raise openai.error.RateLimitError("RateLimitError: rate limit exceeded")
            # if api_key == None, the default openai.api_key will be used

            if session.natal == {}:
                response = openai.ChatCompletion.create(
                    api_key=api_key, messages=session.messages, **self.args
                )
                check=response.choices[0]["message"]["content"]
                if "longitude" in response.choices[0]["message"]["content"]:
                    user_birth_data = json.loads(re.findall(r'{.*}', response.choices[0]["message"]["content"])[0])
                    # ex="session.add_reply(user_yuliao)"
                    session.new()
                    session.natal, session.natal_titles = xingpan(user_birth_data)
                    session.add_reply(session.natal_titles)
                    #session.add_reply('现在我已经获得您的星盘，请问您想占星什么问题？')
                    return {
                        "total_tokens": len(session.messages[-1]["content"]),
                        "completion_tokens": len(session.messages[-1]["content"]),
                        "content": '现在我已经获得您的星盘，请问您想占星什么问题？'
                    }
                # logger.info("[ChatGPT] reply={}, total_tokens={}".format(response.choices[0]['message']['content'], response["usage"]["total_tokens"]))
                session.reset()
                return {
                    "total_tokens": response["usage"]["total_tokens"],
                    "completion_tokens": response["usage"]["completion_tokens"],
                    "content": '请先一次性输入您的出生日期，出生地（城市）和出生时间用以获取您的星盘，例如：1990年12月31日，18点，武汉，如果不确定出生时间可以不输入，将按照中午12点为基准来生成星盘'
                }


            strategy = openpyxl.load_workbook('bot/chatgpt/strategy.xlsx')
            sheets = strategy.worksheets
            sheet_one = sheets[0]
            sheet_two = sheets[1]
            user_question = session.messages[-1]['content']
            solution = ''
            # return {
            #     "total_tokens": len(str(user_question)),
            #     "completion_tokens": len(str(user_question)),
            #     "content": str(user_question)
            # } 测试星盘是否正确获取
            question_list=[]
            solution_list=[]
            for cell in list(sheet_one.columns)[0]:
                question_list.append(cell.value)
            for cell in list(sheet_two.columns)[0]:
                solution_list.append(cell.value)
            # return {
            #     "total_tokens": len(str(user_question)),
            #     "completion_tokens": len(str(user_question)),
            #     "content": str(question_list)
            # } 测试能否正确读取问题列表
            # category = openai.ChatCompletion.create(
            #     api_key=api_key,
            #     model="gpt-3.5-turbo",
            #     temperature=0.7,
            #     max_tokens=512,
            #     messages=[
            #         {"role": "system", "content": "You are a helpful assistant"},
            #         {"role": "assistant", "content": "以下哪个与我输入的意思最接近：" + str(
            #             session.question_list) + '，返回在数组中的下标'},
            #         {"role": "user", "content": str(user_question)}
            #     ]
            # )
            # return {
            #     "total_tokens": len(category.choices[0]["message"]["content"]),
            #     "completion_tokens": len(category.choices[0]["message"]["content"]),
            #     "content": category.choices[0]["message"]["content"]
            # } #测试是否正确分类

            # index= int(re.sub("\D","",category.choices[0]["message"]["content"]))
            # #index = re.sub("\D", "", category.choices[0]["message"]["content"])
            # if len(index) > 0:
            #     solution = ''
            #     index = int(index)
            # for i in range(len(question_list)):
            #     if question_list[i] in category.choices[0]["message"]["content"]:
            titles = []
            question_embedding = openai.Embedding.create(input=user_question, model="text-similarity-davinci-001")['data'][0]['embedding']
            index = 0
            similarity = 0
            for j in range(len(question_list)):
                potential_question_embedding = openai.Embedding.create(input=question_list[j],
                                                                       model="text-similarity-davinci-001")['data'][0]['embedding']
                potential_similarity = np.dot(question_embedding, potential_question_embedding)
                if potential_similarity > 0.81 and potential_similarity > similarity:
                    similarity = potential_similarity
                    index = j+1
            if similarity > 0:
                for i in (2, 5):
                    if sheet_one.cell(index, i).value is not None:
                        try:
                            exec(sheet_one.cell(index, i).value)
                        except:
                            pass

                for t in titles:
                    for j in range(len(solution_list)):
                        if set(t) == set(solution_list[j]):
                            solution += sheet_two.cell(j + 1, 2).value
            strategy.close()
            if solution == '':
                #session.new()
                #session.add_reply(session.natal_titles)
                response = openai.ChatCompletion.create(
                    api_key=api_key, messages=session.messages, **self.args
                )
                # response = openai.ChatCompletion.create(
                #     api_key=api_key,
                #     model="gpt-3.5-turbo",
                #     temperature=0.7,
                #     max_tokens=512,
                #     messages=[
                #         {"role": "system",
                #          "content": "你是一个占星师，请根据我的星盘回答问题"},
                #         {"role": "assistant", "content": session.natal_titles},
                #         {"role": "user", "content": user_question}
                #     ])
                return {
                    "total_tokens": response["usage"]["total_tokens"],
                    "completion_tokens": response["usage"]["completion_tokens"],
                    "content": response.choices[0]["message"]["content"],
                }

            rewrite = openai.ChatCompletion.create(
                api_key=api_key,
                model="gpt-3.5-turbo",
                temperature=0.7,
                max_tokens=512,
                messages=[
                    {"role": "system","content": "你是一个占星师，请总结以下内容并简短回答，因为我不懂占星，所以不需要提示占星术语"},
                    {"role": "user", "content": solution}
                ])
            return {
                "total_tokens": len(rewrite.choices[0]["message"]["content"]),
                "completion_tokens": len(rewrite.choices[0]["message"]["content"]),
                "content": rewrite.choices[0]["message"]["content"]}




        except Exception as e:
            need_retry = retry_count < 2
            result = {"completion_tokens": 0, "content": "我现在有点累了，等会再来吧"}
            if isinstance(e, openai.error.RateLimitError):
                logger.warn("[CHATGPT] RateLimitError: {}".format(e))
                result["content"] = "提问太快啦，请休息一下再问我吧"
                if need_retry:
                    time.sleep(20)
            elif isinstance(e, openai.error.Timeout):
                logger.warn("[CHATGPT] Timeout: {}".format(e))
                result["content"] = "我没有收到你的消息"
                if need_retry:
                    time.sleep(5)
            elif isinstance(e, openai.error.APIConnectionError):
                logger.warn("[CHATGPT] APIConnectionError: {}".format(e))
                need_retry = False
                result["content"] = "我连接不到你的网络"
            elif isinstance(e, IndexError):
                logger.warn("Birth data transformation Fault : {}".format(e))
                need_retry = False
                session.reset()
                result["content"] = "读取生日错误，请重试"
            else:
                logger.warn("[CHATGPT] Exception: {}".format(e))
                need_retry = False
                self.sessions.clear_session(session.session_id)

            if need_retry:
                logger.warn("[CHATGPT] 第{}次重试".format(retry_count + 1))
                return self.reply_text(session, api_key, retry_count + 1)
            else:
                return result


class AzureChatGPTBot(ChatGPTBot):
    def __init__(self):
        super().__init__()
        openai.api_type = "azure"
        openai.api_version = "2023-03-15-preview"
        self.args["deployment_id"] = conf().get("azure_deployment_id")
